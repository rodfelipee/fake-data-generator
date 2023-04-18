from faker import Faker
from openpyxl import Workbook

# Cria um objeto Faker
fake = Faker()

# Cria uma planilha do Excel
wb = Workbook()

# Seleciona a primeira planilha
ws = wb.active

# Cria as células para os dados pessoais
ws['A1'] = 'Nome'
ws['B1'] = 'Endereço'
ws['C1'] = 'Telefone'
ws['D1'] = 'E-mail'

# Preenche a planilha com dados pessoais fictícios
for row in range(2, 101):
    nome = fake.name()
    endereco = fake.address()
    telefone = fake.phone_number()
    email = fake.email()
    ws.cell(row=row, column=1, value=nome)
    ws.cell(row=row, column=2, value=endereco)
    ws.cell(row=row, column=3, value=telefone)
    ws.cell(row=row, column=4, value=email)

# Salva a planilha em um arquivo
wb.save('data.xlsx')
